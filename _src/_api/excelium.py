#!/usr/bin/python
# excel.py
# -*- coding: utf-8 -*-

'''
Created on 2019. 1. 14.
@author: miskang
'''

# import requests library for excel call
import openpyxl

# =============================================================================================


'''
class workbook(filename) 
def init -get workbook return wb
def get_sheet_list(): return sheet(list)
def get_worksheet(sheet): return header(list), data(list)
def save_workbook(filename): return 0
def close_workbook(): return 0
'''


class Workbook(object):
    def __init__(self, file,read_only,data_only):
        self.wb = openpyxl.load_workbook(file, read_only=read_only,data_only=data_only)
    
    def create_new_sheet(self,sheet_name):
        self.ws = self.wb.create_sheet(sheet_name)
        return self.ws

    def remove_sheet(self,sheet_name):
        self.ws = self.wb.get_sheet_by_name(sheet_name)
        self.wb.remove(self.ws)
        return 0

    def get_sheet_list(self):
        # print(self.wb.sheetnames)
        return self.wb.sheetnames

    def get_first_row(self, sheet):
        self.ws = self.wb.get_sheet_by_name(sheet)
        self.first_row = self.ws[1]
        self.index = []
        for r in self.first_row:
            self.index.append(r.value)
        return self.index

    def get_worksheet(self, sheet_name):
        self.ws = self.wb.get_sheet_by_name(sheet_name)
        return self.ws

    def change_cell_data(self, ws, col, row, val):
        # set value for each cell
        # col = A,B,C.... row = 1,2,3....
        ws.cell(row=row, column=col, value=val)
        # ws[str(cell)] = val
        return 0

    def save_workbook(self, file):
        try:
            self.wb.save(file)
        except PermissionError:
            self.wb.save(str(file).replace('.xlsx','_temp.xlsx'))
        return 0

    def close_workbook(self):
        self.wb.close()
        return 0
